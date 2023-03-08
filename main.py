import requests
import json
import openpyxl

part_file = '/Users/bubulindo/Documents/Book2.xlsx' # type your path to your parts file.
results_file = '/Users/bubulindo/Documents/parts.xlsx' # type your path to the results file to be created.


# This function merely returns the beautified json response from Rockwell's servers to allow for more
# functionality in the script below.
def get_status(part_number):
    url = "https://es-be-ux-search.cloudhub.io/api/ux/v2/search?queryText=" + part_number + "&role=rockwell-search&spellingCorrect=true&spellcheckPremium=10&segments=Productsv4&startIndex=0&numResults=20&facets=&languages=en&locales=en-GB,en_GLOBAL&sort=bma&collections=Literature,Web,Sample_Code&site=RA"
    headers = {'client_id': 'fb000cbbe476420b9e70be741abd7a63', 'client_secret': 'Db420ae8BAdD47ADA4E12cE90Fb1b747', 'correlation_id': '1eaaba80-16be-6c60-9dd6-1378852fe624'}

    # client_secret
    # this value was pulled from the web browser with the "code inspector" and is required for the request to come through.
    # Db420ae8BAdD47ADA4E12cE90Fb1b747
    response = requests.request("GET", url, headers=headers)
    if response.status_code != 200:
        return "no network" # check your wifi and try again.
    # print (response)  # Troubleshooting
    responseData = json.loads(response.content)
    return responseData

# Start of the script.

# open workbook where the part numbers are.
wb = openpyxl.load_workbook(filename=part_file)

# Open the Sheet where the data is stored.
sheet = wb['Sheet1']

# Create a new workbook and sheet for the results to be stored.
respBook = openpyxl.Workbook() # create a new workbook.
respSheet = respBook.create_sheet("Sheet1")  # create a new sheet in the response workbook.
# print(sheet.cell(column=1, row =10).value)  #Troubleshooting

# the variable below keeps track of the line in the new file to be written as we go up and down on counters.
# This is the most straightforward way, even if not very pretty.
responseLine = 2 # this keeps track of where to put the response

#create result header in the new workbook sheet.
respSheet.cell(column=1, row=1).value = "Part Number"
respSheet.cell(column=2, row=1).value = "Technical Description"
respSheet.cell(column=3, row=1).value = "Lifecycle Status"
respSheet.cell(column=4, row=1).value = "Repairable"
respSheet.cell(column=5, row=1).value = "Discontinued Date"

# Cycle through the entries in the parts list. This is capped at 200 entries as it should be enough for parts.
# If the cycle detects an empty cell, it will jump out, so no worry about being too many entries.
# if required, adjust the 200 figure up to catch all entries.
for row in range(2, 200):

    print(row)  # Troubleshooting

# when it finds an empty row, it breaks the cycle
    if sheet.cell(column=1, row=row).value is None:
        break

# dados will keep the JSON response from Rockwell's servers.
    dados = get_status(sheet.cell(column=1, row=row).value)
# Let's fill in the sheet cell with the part number that we are looking for in case it isn't found.
    # if the part is found, this cell will be overwritten below.
    respSheet.cell(column=1, row=responseLine).value = sheet.cell(column=1, row=row).value # part number.
    # check if any records were retrieved
    if dados['response']['numFound'] == 0:
        respSheet.cell(column=2, row=responseLine).value = "not found" # no records retrieved from Rockwell.
    else:  # there's data. Some records were found.
        numRecords = dados['response']['numFound'] # this is the number of records found
        # I limit the records to 30 as it's been enough so far. If you look in the url variable with the Rockwell's website
        # you'll see that there's also a hard limit there of 30. This can create problems as their site will only return
        # up to the number of records provided... so it can state that it has found 100 records but only sends out 30.
        if numRecords > 30:
            numRecords = 30
        # this is the starting point of the records... just being precious.
        startRecords = dados['response']['start']
        print(dados)  # Troubleshooting
        # run through the records that were received.
        # for each record a line in the Excel sheet will have part number (overwritten), technical Description,
        # lifecycle Status, discontinued date and Repairable (these two only if sent by Rockwell).
        for inner in range(startRecords, numRecords):
            respSheet.cell(column=1, row=responseLine).value = dados['response']['docs'][inner]['catalogNumber']  # part number.
            respSheet.cell(column=2, row=responseLine).value = dados['response']['docs'][inner]['technicalDescription']
            respSheet.cell(column=3, row=responseLine).value = dados['response']['docs'][inner]['lifecycleStatus']
            # these don't show up all the time, so it has to be on a try block so it doesn't crash the script.
            try:
                tempDate = dados['response']['docs'][inner]['discontinuedDate']
                tempRepair = dados['response']['docs'][inner]['repairable']
            except:  # if there's an exception, then the excel line will show not set as per below.
                tempRepair = "not set"
                tempDate = "not set"
            respSheet.cell(column=4, row=responseLine).value = tempRepair
            respSheet.cell(column=5, row=responseLine).value = tempDate
            # increment the line that is to be written next in the Excel file.
            responseLine = responseLine + 1 # increment
            print("--->" + str(responseLine))  # Troubleshooting

# save all of this faffing in a new EXcel file. This script needs to be run on my Macbook or the path below changed.
respBook.save(filename=results_file)

# record all the records. This can then be filtered in the Excel document


